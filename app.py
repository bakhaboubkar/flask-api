# -*- coding: utf-8 -*-
"""FC SCAN — Flask API Backend — نفس منطق model.py + historique.py"""

import os, sys, io
from datetime import date
import numpy as np
import joblib
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

# Fix NumPy aliases
for _a, _t in [('float', float),('int', int),('complex', complex),
               ('bool', bool),('object', object),('str', str)]:
    if not hasattr(np, _a): setattr(np, _a, _t)

import pyrenn as pr

app = Flask(__name__)
CORS(app)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Load model ────────────────────────────────────────────────────────────────
net = scaler_X = scaler_y = None
LM_OK = False
try:
    net      = pr.loadNN(os.path.join(SCRIPT_DIR, "ann_regression_model_pyrenn.csv"))
    scaler_X = joblib.load(os.path.join(SCRIPT_DIR, "scaler_X.pkl"))
    scaler_y = joblib.load(os.path.join(SCRIPT_DIR, "scaler_y.pkl"))
    LM_OK = True
    print(f"[FC SCAN] Modèle ANN-LM chargé depuis: {SCRIPT_DIR}")
except Exception as ex:
    print(f"[FC SCAN] ERREUR: {ex}")

# ── Concrete classes — نفس CONCRETE_CLASSES ──────────────────────────────────
CONCRETE_CLASSES = [
    (0,   16,  "#E74C3C", "Très faible", "C16/20"),
    (16,  20,  "#E67E22", "Faible",      "C20/25"),
    (20,  25,  "#F39C12", "Ordinaire",   "C25/30"),
    (25,  30,  "#2ECC71", "Résistant",   "C30/37"),
    (30, 999,  "#2E86AB", "Haute perf.", "C35/45"),
]

CONF_FILLS = {
    "Très faible": "FEF0EF",
    "Faible":      "FEF3EC",
    "Ordinaire":   "FEF9EC",
    "Résistant":   "EDFAF3",
    "Haute perf.": "EAF5FB",
}

def get_class(fc):
    for i,(lo,hi,color,label,ec) in enumerate(CONCRETE_CLASSES):
        if lo <= fc < hi:
            return {"label":label,"color":color,"ec":ec,"index":i}
    c = CONCRETE_CLASSES[-1]
    return {"label":c[3],"color":c[2],"ec":c[4],"index":len(CONCRETE_CLASSES)-1}

def predict_model(R, V):
    if not LM_OK: return {"error":"Aucun modèle chargé"}
    try:
        x  = np.array([R,V]).reshape(1,-1)
        xs = scaler_X.transform(x)
        ys = pr.NNOut(xs.T, net)
        fc = float(scaler_y.inverse_transform(np.array(ys).reshape(-1,1))[0,0])
        return {"fc":fc,"detail":f"ANN-LM fc={fc:.4f} MPa","model":"ANN — LM"}
    except Exception as ex:
        return {"error":str(ex)}

# ═════════════════════════════════════════════════════════════════
# ROUTES
# ═════════════════════════════════════════════════════════════════

@app.route("/health")
def health():
    return jsonify({"status":"ok","model_loaded":LM_OK,
                    "model":"ANN — LM (pyrenn)" if LM_OK else None})

@app.route("/predict", methods=["POST"])
def predict():
    d = request.get_json(force=True)
    try: R,V = float(d["R"]), float(d["V"])
    except: return jsonify({"error":"R et V requis"}), 400
    res = predict_model(R,V)
    if "error" in res: return jsonify(res), 500
    res["class"] = get_class(res["fc"])
    return jsonify(res)

@app.route("/predict_batch", methods=["POST"])
def predict_batch():
    d = request.get_json(force=True)
    results = []
    for item in d.get("items",[]):
        try: R,V = float(item["R"]), float(item["V"])
        except: results.append({"error":"R/V invalide","element":item.get("element","")});continue
        res = predict_model(R,V)
        if "error" in res: results.append({**res,"element":item.get("element","")});continue
        res["class"]   = get_class(res["fc"])
        res["element"] = item.get("element","")
        fcd = item.get("fcd",0)
        if fcd and float(fcd)>0:
            res["conforme"] = res["fc"] >= float(fcd)
            res["fcd"]      = float(fcd)
        results.append(res)
    return jsonify({"results":results})

# ── Export Excel — نفس _export_excel() في historique.py ─────────────────────
@app.route("/export/excel", methods=["POST"])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except:
        return jsonify({"error":"openpyxl requis: pip install openpyxl"}), 500

    d       = request.get_json(force=True)
    project = d.get("project",{})
    points  = d.get("points",[])
    stats   = d.get("stats",{})

    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    PROJ_FILL = PatternFill("solid", fgColor="EEF2F7")
    BOLD      = Font(bold=True, size=9)
    NORMAL    = Font(size=9)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    thin      = Side(style="thin", color="D8E2ED")
    BDR       = Border(left=thin, right=thin, top=thin, bottom=thin)
    CONF_FILLS_XL = {k: PatternFill("solid", fgColor=v) for k,v in CONF_FILLS.items()}

    wb  = openpyxl.Workbook()

    # Sheet 1 — Projet
    ws1 = wb.active; ws1.title = "Projet"
    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 36
    ws1.merge_cells("A1:B1")
    c = ws1["A1"]; c.value = "FC SCAN — Rapport d'Inspection"
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.fill = HDR_FILL; c.alignment = CENTER; ws1.row_dimensions[1].height = 28

    info_rows = [
        ("Nom du projet",      project.get("project_name","")),
        ("Structure",          project.get("structure_name","")),
        ("Localisation",       project.get("location","")),
        ("Inspecteur",         project.get("inspector","")),
        ("Date d'inspection",  project.get("inspection_date","")),
        ("Modèle ANN",         "ANN — Levenberg-Marquardt (pyrenn)"),
        ("Points enregistrés", str(len(points))),
    ]
    for r,(k,v) in enumerate(info_rows, 2):
        ws1[f"A{r}"].value=k; ws1[f"A{r}"].font=BOLD
        ws1[f"A{r}"].fill=PROJ_FILL; ws1[f"A{r}"].border=BDR; ws1[f"A{r}"].alignment=LEFT
        ws1[f"B{r}"].value=v; ws1[f"B{r}"].font=NORMAL
        ws1[f"B{r}"].border=BDR; ws1[f"B{r}"].alignment=LEFT

    # Sheet 2 — Mesures
    ws2 = wb.create_sheet("Mesures")
    headers = ["Élément","R","V (µm/s)","fc (MPa)","Classe"]
    col_ws  = [24,8,12,12,16]
    for ci,(h,cw) in enumerate(zip(headers,col_ws),1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width=cw
        c=ws2.cell(row=1,column=ci,value=h)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=BDR
    for ri,pt in enumerate(points,2):
        vals=[pt.get("element",""), pt.get("R",0), pt.get("V",0),
              round(pt.get("fc",0),2), pt.get("confidence","")]
        cf=CONF_FILLS_XL.get(pt.get("confidence",""))
        for ci,val in enumerate(vals,1):
            c=ws2.cell(row=ri,column=ci,value=val)
            c.font=NORMAL; c.alignment=CENTER; c.border=BDR
            if cf: c.fill=cf

    # Sheet 3 — Statistiques
    ws3 = wb.create_sheet("Statistiques")
    ws3.column_dimensions["A"].width=28; ws3.column_dimensions["B"].width=18
    ws3.merge_cells("A1:B1"); c=ws3["A1"]
    c.value="Statistiques fc (MPa)"; c.font=HDR_FONT
    c.fill=HDR_FILL; c.alignment=CENTER; ws3.row_dimensions[1].height=24
    for r,(k,v) in enumerate([
        ("Nombre de points", stats.get("count",0)),
        ("Moyenne (MPa)",    stats.get("mean",0)),
        ("Minimum (MPa)",    stats.get("min",0)),
        ("Maximum (MPa)",    stats.get("max",0)),
    ],2):
        ws3[f"A{r}"].value=k; ws3[f"A{r}"].font=BOLD
        ws3[f"A{r}"].fill=PROJ_FILL; ws3[f"A{r}"].border=BDR; ws3[f"A{r}"].alignment=LEFT
        ws3[f"B{r}"].value=v; ws3[f"B{r}"].font=NORMAL
        ws3[f"B{r}"].border=BDR; ws3[f"B{r}"].alignment=CENTER

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    proj_name = project.get("project_name","FC_Scan") or "FC_Scan"
    fname = f"{proj_name}_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Export PDF — نفس _export_pdf() في historique.py ─────────────────────────
@app.route("/export/pdf", methods=["POST"])
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except:
        return jsonify({"error":"reportlab requis: pip install reportlab"}), 500

    d       = request.get_json(force=True)
    project = d.get("project",{})
    points  = d.get("points",[])
    stats   = d.get("stats",{})

    HDR_COLOR = colors.HexColor("#1E3A5F")
    CONF_COLORS_PDF = {
        "Très faible": colors.HexColor("#FEF0EF"),
        "Faible":      colors.HexColor("#FEF3EC"),
        "Ordinaire":   colors.HexColor("#FEF9EC"),
        "Résistant":   colors.HexColor("#EDFAF3"),
        "Haute perf.": colors.HexColor("#EAF5FB"),
    }

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TFC", parent=styles["Title"],
                                  textColor=HDR_COLOR, alignment=TA_CENTER, fontSize=16)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             topMargin=18*mm, bottomMargin=18*mm,
                             leftMargin=15*mm, rightMargin=15*mm)
    elems = []
    elems.append(Paragraph("FC SCAN — Rapport d'Inspection", title_style))
    elems.append(Spacer(1,4))
    line_tbl = Table([[""]], colWidths=[doc.width])
    line_tbl.setStyle(TableStyle([("LINEBELOW",(0,0),(-1,-1),1.2,HDR_COLOR)]))
    elems.append(line_tbl)
    elems.append(Spacer(1,10))

    proj_rows = [
        ("Nom du projet",       project.get("project_name","")),
        ("Nom de la structure", project.get("structure_name","")),
        ("Localisation",        project.get("location","")),
        ("Inspecteur",          project.get("inspector","")),
        ("Date d'inspection",   project.get("inspection_date","")),
    ]
    half = (len(proj_rows)+1)//2
    left_r,right_r = proj_rows[:half], proj_rows[half:]
    while len(left_r)<len(right_r): left_r.append(("",""))
    while len(right_r)<len(left_r): right_r.append(("",""))
    info_data = [[lk,lv,rk,rv] for (lk,lv),(rk,rv) in zip(left_r,right_r)]
    info_tbl = Table(info_data, colWidths=[35*mm,57*mm,35*mm,53*mm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
        ("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("BACKGROUND",(0,0),(0,-1),colors.HexColor("#EEF2F7")),
        ("BACKGROUND",(2,0),(2,-1),colors.HexColor("#EEF2F7")),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#D8E2ED")),
        ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
    ]))
    elems.append(info_tbl)
    elems.append(Spacer(1,14))

    elems.append(Paragraph("Tableau des mesures", ParagraphStyle(
        "H2",parent=styles["Heading2"],textColor=HDR_COLOR,fontSize=12)))
    elems.append(Spacer(1,4))
    headers = ["Élément","R","V (µm/s)","fc (MPa)","Classe"]
    data = [headers]
    row_colors = [None]
    for pt in points:
        data.append([pt.get("element",""), pt.get("R",0), pt.get("V",0),
                     round(pt.get("fc",0),2), pt.get("confidence","")])
        row_colors.append(CONF_COLORS_PDF.get(pt.get("confidence","")))
    meas_tbl = Table(data, colWidths=[60*mm,22*mm,30*mm,30*mm,38*mm], repeatRows=1)
    meas_style = [
        ("BACKGROUND",(0,0),(-1,0),HDR_COLOR),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8.5),
        ("ALIGN",(1,0),(-1,-1),"CENTER"),
        ("ALIGN",(0,0),(0,-1),"LEFT"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#D8E2ED")),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
    ]
    for ri,cf in enumerate(row_colors):
        if cf: meas_style.append(("BACKGROUND",(0,ri),(-1,ri),cf))
    meas_tbl.setStyle(TableStyle(meas_style))
    elems.append(meas_tbl)
    elems.append(Spacer(1,14))

    elems.append(Paragraph("Statistiques fc (MPa)", ParagraphStyle(
        "H2b",parent=styles["Heading2"],textColor=HDR_COLOR,fontSize=12)))
    elems.append(Spacer(1,4))
    stats_tbl = Table(
        [["n","Moyenne (MPa)","Min (MPa)","Max (MPa)"],
         [stats.get("count",0),stats.get("mean",0),
          stats.get("min",0),stats.get("max",0)]],
        colWidths=[45*mm,45*mm,45*mm,45*mm])
    stats_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),HDR_COLOR),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),9),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#D8E2ED")),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
    ]))
    elems.append(stats_tbl)
    elems.append(Spacer(1,20))
    elems.append(Paragraph(
        f"FC SCAN v1.0  •  Évaluation du Béton par Analyse Numérique  •  "
        f"{date.today().strftime('%d/%m/%Y')}",
        ParagraphStyle("Footer",parent=styles["Normal"],fontSize=8,
                       textColor=colors.HexColor("#888888"),alignment=TA_CENTER)))
    doc.build(elems)
    buf.seek(0)
    proj_name = project.get("project_name","FC_Scan") or "FC_Scan"
    fname = f"{proj_name}_{date.today().strftime('%Y%m%d')}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/pdf")

# ── Import Excel ──────────────────────────────────────────────────────────────
@app.route("/import/excel", methods=["POST"])
def import_excel():
    import re, unicodedata
    try: import openpyxl
    except: return jsonify({"error":"openpyxl requis"}),500

    if "file" not in request.files:
        return jsonify({"error":"Fichier manquant"}),400
    f = request.files["file"]
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as ex:
        return jsonify({"error":str(ex)}),400

    if len(all_rows)<2:
        return jsonify({"error":"Fichier vide ou sans données"}),400

    def norm(h):
        if h is None: return ""
        s=str(h).strip().lower()
        s=unicodedata.normalize("NFD",s)
        s="".join(c for c in s if unicodedata.category(c)!="Mn")
        s=re.sub(r"[^a-z0-9 ]"," ",s)
        return re.sub(r"\s+"," ",s).strip()

    R_AL={"r","rebound","schmidt","indice r","indice sclerometrique","ir","r value","rebond"}
    V_AL={"v","upv","vitesse","velocity","pulse velocity","v m s","v um s","ultrasonic","vp"}
    E_AL={"element","element id","elementid","id","repere","reference","label","name","nom"}

    raw_hdrs = list(all_rows[0])
    r_col=v_col=e_col=None
    for i,h in enumerate(raw_hdrs):
        n=norm(h)
        if r_col is None and n in R_AL: r_col=i
        elif v_col is None and n in V_AL: v_col=i
        elif e_col is None and n in E_AL: e_col=i

    # Fallback by value range
    if r_col is None or v_col is None:
        col_means=[]
        for ci in range(len(raw_hdrs)):
            vals=[float(r[ci]) for r in all_rows[1:21] if ci<len(r) and r[ci] is not None
                  and str(r[ci]).replace('.','').replace('-','').isdigit()]
            col_means.append(sum(vals)/len(vals) if vals else None)
        for ci,m in enumerate(col_means):
            if m is None: continue
            if r_col is None and 5<=m<=80: r_col=ci
            elif v_col is None and (200<=m<=8000 or 0.2<=m<=8.0): v_col=ci

    if r_col is None or v_col is None:
        return jsonify({"error":"Colonnes R et/ou V non détectées"}),400

    imported,rejected,results=[],[],[]
    counter=0
    for row in all_rows[1:]:
        row=list(row)
        counter+=1
        elem=str(row[e_col]).strip() if e_col is not None and e_col<len(row) and row[e_col] else f"P{counter}"
        try: R=float(row[r_col])
        except: rejected.append({"row":counter,"reason":f"R invalide"}); continue
        try:
            raw_v=float(row[v_col])
            V=raw_v*1000 if 0.2<=raw_v<=8.0 else raw_v
        except: rejected.append({"row":counter,"reason":f"V invalide"}); continue
        if not(5<=R<=80): rejected.append({"row":counter,"reason":f"R={R} hors plage"}); continue
        if not(200<=V<=8000): rejected.append({"row":counter,"reason":f"V={V} hors plage"}); continue
        res=predict_model(R,V)
        if "error" in res: rejected.append({"row":counter,"reason":res["error"]}); continue
        fc=res["fc"]; cls=get_class(fc)
        imported.append({"element":elem,"R":R,"V":V,"fc":fc,
                         "confidence":cls["label"],"color":cls["color"],
                         "ec":cls["ec"],"pointId":f"P{counter}"})

    return jsonify({"imported":imported,"rejected":rejected,
                    "total":len(all_rows)-1,
                    "imported_count":len(imported),"rejected_count":len(rejected)})

if __name__ == "__main__":
    port = int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0", port=port, debug=False)
